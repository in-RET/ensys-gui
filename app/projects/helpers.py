import csv
import io
import json
import os
import re

import scipy
from InRetEnsys import *
from dashboard.helpers import KPIFinder
from django import forms
from django.contrib.staticfiles.storage import staticfiles_storage
from django.core.exceptions import ValidationError
from django.utils.translation import gettext_lazy as _
from oemof.tools import economics
from openpyxl import load_workbook
from projects.constants import MAP_MVS_EPA
from projects.dtos import convert_to_dto
from projects.models import InputparameterSuggestion, Bus

PARAMETERS = {}
if os.path.exists(staticfiles_storage.path("MVS_parameters_list.csv")) is True:
    with open(
        staticfiles_storage.path("MVS_parameters_list.csv"), encoding="utf-8"
    ) as csvfile:
        csvreader = csv.reader(csvfile, delimiter=",", quotechar='"')
        for i, row in enumerate(csvreader):
            if i == 0:
                hdr = row
                label_idx = hdr.index("label")
            else:
                label = row[label_idx]
                label = MAP_MVS_EPA.get(label, label)
                PARAMETERS[label] = {}
                for k, v in zip(hdr, row):
                    if k == "sensitivity_analysis":
                        v = bool(int(v))
                    PARAMETERS[label][k] = v

parameters_helper = KPIFinder(param_info_dict=PARAMETERS, unit_header=":Unit:")


# Helper method to clean dict data from empty values
def remove_empty_elements(d):
    def empty(x):
        return x is None or x == {} or x == []

    if not isinstance(d, (dict, list)):
        return d
    elif isinstance(d, list):
        return [v for v in (remove_empty_elements(v) for v in d) if not empty(v)]
    else:
        return {
            k: v
            for k, v in ((k, remove_empty_elements(v)) for k, v in d.items())
            if not empty(v)
        }


# Helper to convert Scenario data to MVS importable json
def format_scenario_for_mvs(scenario_to_convert, testing=False):
    mvs_request_dto = convert_to_dto(scenario_to_convert, testing=testing)
    # print(mvs_request_dto)
    dumped_data = json.loads(
        json.dumps(mvs_request_dto.__dict__, default=lambda o: o.__dict__)
    )

    # format the constraints in MVS format directly, thus avoiding the need to maintain MVS-EPA
    # parser in multi-vector-simulator package
    constraint_dict = {}
    for constraint in dumped_data["constraints"]:
        constraint_dict[constraint["label"]] = constraint["value"]
    dumped_data["constraints"] = constraint_dict

    # Remove None values
    return dumped_data  # remove_empty_elements


def epc_calc(capex, Amortisierungszeit, opex, interest_rate):
    if interest_rate == 0:
        interest_rate = 1 * 10**-4  # to avoid division by zero error
    investk = economics.annuity(
        capex=capex, n=Amortisierungszeit, wacc=interest_rate / 100
    )
    betriebsk = capex * (opex / 100)
    epc = investk + betriebsk
    return epc


def polate_unknown_capex(technology, year, asset_type_name):
    queryset_2030 = InputparameterSuggestion.objects.filter(
        technology=technology, year=2030
    )
    queryset_2040 = InputparameterSuggestion.objects.filter(
        technology=technology, year=2040
    )

    for item in queryset_2030:
        # print(item.unique_id, item.capex)
        capex_2030 = item.capex
        opex = item.opex
        lifetime = item.lifetime
        efficiency = item.efficiency
        efficiency_el = item.efficiency_el
        efficiency_th = item.efficiency_th
        input_timeseries = item.input_timeseries
        thermal_loss_rate = item.thermal_loss_rate
        fixed_losses_relative_gamma = item.fixed_losses_relative_gamma
        fixed_losses_absolute_delta = item.fixed_losses_absolute_delta
        crate = item.crate
        # print(crate)

    for item in queryset_2040:
        # print(item.unique_id, item.capex)
        capex_2040 = item.capex

    x = [2030, 2040]
    y = [capex_2030, capex_2040]

    if year == "2025":
        f_ex = scipy.interpolate.interp1d(x, y, fill_value="extrapolate")
        capex_new = f_ex(2025)
        # print(capex_new, capex_2030, capex_2040)

    elif year == "2035":
        f_linear = scipy.interpolate.interp1d(x, y)
        capex_new = f_linear(2035)
        # print(capex_2030, capex_new, capex_2040)

    if asset_type_name == "myPredefinedStorage":
        return (
            capex_new,
            opex,
            lifetime,
            crate,
            efficiency,
            thermal_loss_rate,
            fixed_losses_relative_gamma,
            fixed_losses_absolute_delta,
        )

    elif asset_type_name == "myPredefinedTransformer":
        return (
            capex_new,
            opex,
            lifetime,
            efficiency,
            efficiency_el,
            efficiency_th,
            input_timeseries,
        )

    elif asset_type_name == "myPredefinedSource":
        return capex_new, opex, lifetime, input_timeseries


def sensitivity_analysis_payload(
    variable_parameter_name="",
    variable_parameter_range="",
    variable_parameter_ref_val="",
    output_parameter_names=None,
):
    """format the parameters required to request a sensitivity analysis in a specific JSON"""
    if output_parameter_names is None:
        output_parameter_names = []
    return {
        "sensitivity_analysis_settings": {
            "variable_parameter_name": variable_parameter_name,
            "variable_parameter_range": variable_parameter_range,
            "variable_parameter_ref_val": variable_parameter_ref_val,
            "output_parameter_names": output_parameter_names,
        }
    }


SA_RESPONSE_SCHEMA = {
    "type": "object",
    "required": ["server_info", "mvs_version", "id", "status", "results"],
    "properties": {
        "server_info": {"type": "string"},
        "mvs_version": {"type": "string"},
        "id": {"type": "string"},
        "status": {"type": "string"},
        "results": {
            "type": "object",
            "required": ["reference_simulation_id", "sensitivity_analysis_steps"],
            "properties": {
                "reference_simulation_id": {"type": "string"},
                "sensitivity_analysis_steps": {
                    "type": "array",
                    "items": {"type": "object"},
                },
            },
            "additionalProperties": False,
        },
        "ref_sim_id": {"type": "string"},
        "sensitivity_analysis_ids": {"type": "array", "items": {"type": "string"}},
    },
    "additionalProperties": False,
}


# Used to proof the json objects stored as text in the db
SA_OUPUT_NAMES_SCHEMA = {"type": "array", "items": {"type": "string"}}


def sa_output_values_schema_generator(output_names):
    return {
        "type": "object",
        "required": output_names,
        "properties": {
            output_name: {
                "type": "object",
                "required": ["value", "path"],
                "properties": {
                    "value": {
                        "oneOf": [
                            {"type": "null"},
                            {
                                "type": "array",
                                "items": {
                                    "anyOf": [{"type": "number"}, {"type": "null"}]
                                },
                            },
                        ]
                    },
                    "path": {
                        "oneOf": [
                            {"type": "string"},
                            {"type": "array", "items": {"type": "string"}},
                        ]
                    },
                },
            }
            for output_name in output_names
        },
        "additionalProperties": False,
    }


class DualInputWidget(forms.MultiWidget):

    template_name = "asset/dual_input.html"

    def __init__(self, **kwargs):
        """This special input consist of one text field and one upload file button"""

        self.default = kwargs.pop("default", None)
        self.param_name = kwargs.pop("param_name", None)

        widgets = {
            "scalar": forms.TextInput(
                attrs={
                    "class": "form-control",
                    "onchange": f"plotDualInputTrace(obj=this.value, param_name='{self.param_name}')",
                }
            ),
            "file": forms.FileInput(
                attrs={
                    "class": "form-control",
                    "onchange": f"uploadDualInputTrace(obj=this.files, param_name='{self.param_name}')",
                }
            ),
        }
        super(DualInputWidget, self).__init__(widgets=widgets, **kwargs)

    def use_required_attribute(self, initial):
        # overwrite the method of the Widget class of the django.form.widgets module
        return False

    def decompress(self, value):
        answer = [self.default, None]
        if value is not None:
            value = json.loads(value)
            answer = [value, None]
        return answer


class DualNumberField(forms.MultiValueField):
    def __init__(self, default=None, param_name=None, **kwargs):
        fields = (forms.DecimalField(required=False), forms.CharField(required=False))
        kwargs.pop("max_length", None)
        self.min = kwargs.pop("min", None)
        self.max = kwargs.pop("max", None)
        kwargs["widget"] = DualInputWidget(default=default, param_name=param_name)

        super().__init__(fields=fields, require_all_fields=False, **kwargs)

    def clean(self, values):
        """If a file is provided it will be considered over the scalar"""
        scalar_value, timeseries_file = values

        if timeseries_file is not None:
            input_timeseries_values = parse_input_timeseries(timeseries_file)
            answer = input_timeseries_values
        else:

            # check the input string is a number or a list
            if scalar_value != "":
                try:
                    answer = float(scalar_value)
                except ValueError:
                    try:
                        answer = json.loads(scalar_value)
                        if not isinstance(answer, list):
                            scalar_value = ""
                    except json.decoder.JSONDecodeError:
                        scalar_value = ""

            if scalar_value == "":
                self.set_widget_error()
                raise ValidationError(
                    _(
                        "Please provide either a number within %(boundaries) s or upload a timeseries from a file"
                    ),
                    code="required",
                    params={"boundaries": self.boundaries},
                )
        self.check_boundaries(answer)
        return answer

    @property
    def boundaries(self):
        if self.min is not None:
            min_val = self.min
        else:
            min_val = "-inf"

        if self.max is not None:
            max_val = self.max
        else:
            max_val = "inf"

        return f"[{min_val}, {max_val}]"

    def check_boundaries(self, value):

        boundaries = self.boundaries
        if isinstance(value, list):
            for v in value:
                try:
                    self.check_boundaries(v)
                except ValidationError:
                    self.set_widget_error()
                    raise ValidationError(
                        _(
                            "Some values in the timeseries do not lie within %(boundaries) s, please check your input again."
                        ),
                        code="invalid",
                        params={"boundaries": boundaries},
                    )

        else:
            if self.min is not None:
                if value < self.min:
                    self.set_widget_error()
                    raise ValidationError(
                        _("Please enter a value within %(boundaries) s"),
                        code="invalid",
                        params={"boundaries": boundaries},
                    )

            if self.max is not None:
                if value > self.max:
                    self.set_widget_error()
                    raise ValidationError(
                        _("Please enter a value within %(boundaries) s"),
                        code="invalid",
                        params={"boundaries": boundaries},
                    )

    def set_widget_error(self):
        for widget in self.widget.widgets:
            css = widget.attrs.get("class", None)
            if css is not None:
                css = css.split(" ")
            else:
                css = []
            css.append("is-invalid")
            widget.attrs["class"] = " ".join(css)


def parse_csv_timeseries(file_str):
    io_string = io.StringIO(file_str)
    delimiter = ","
    if file_str.count(";") > 0:
        delimiter = ";"

    # check if the number of , is an integer time the number of line return
    # if not, the , is probably not a column separator and a decimal separator indeed
    if file_str.count(",") % (file_str.count("\n") + 1) != 0:
        delimiter = ";"

    reader = csv.reader(io_string, delimiter=delimiter)
    timeseries_values = []
    for row in reader:
        if len(row) == 1:
            value = row[0]
        else:
            # assumes the first row is timestamps and read the second one, ignore any other row
            value = row[1]
        # convert potential comma used as decimal point to decimal point
        timeseries_values.append(float(value.replace(",", ".")))
    return timeseries_values


def parse_input_timeseries(timeseries_file):
    if timeseries_file.name.endswith("xls") or timeseries_file.name.endswith("xlsx"):
        wb = load_workbook(filename=timeseries_file)
        worksheet = wb.active
        timeseries_values = []
        n_col = worksheet.max_column

        col_idx = 0

        if n_col > 1:
            col_idx = 1

        for j in range(0, worksheet.max_row):
            try:
                timeseries_values.append(
                    float(worksheet.cell(row=j + 1, column=col_idx + 1).value)
                )
            except ValueError:
                pass

    else:
        timeseries_file_str = timeseries_file.read().decode("utf-8")

        if timeseries_file_str != "":
            if timeseries_file.name.endswith("json"):
                timeseries_values = json.loads(timeseries_file_str)
            elif timeseries_file.name.endswith("csv"):
                timeseries_values = parse_csv_timeseries(timeseries_file_str)

            elif timeseries_file.name.endswith("txt"):
                nlines = timeseries_file_str.count("\n") + 1
                if nlines == 1:
                    timeseries_values = json.loads(timeseries_file_str)
                else:
                    timeseries_values = parse_csv_timeseries(timeseries_file_str)
            else:
                raise TypeError(
                    _(
                        f'Input timeseries file type of "{timeseries_file.name}" is not supported. The supported formats are "json", "csv", "txt", "xls" and "xlsx"'
                    )
                )
        else:
            raise ValidationError(
                _('Input timeseries file "%(fname)s" is empty'),
                code="empty_file",
                params={"fname": timeseries_file.name},
            )
    return timeseries_values


def expert_trafo_parameter_visibility(form, combination):
    if (
        combination == "2:3"
        or combination == "1:2"
        or combination == "1:1"
        or combination == "2:1"
    ):
        field = form.fields["trafo_input_bus_3"]
        field.widget = field.hidden_widget()
        field = form.fields["trafo_input_conversionf_3"]
        field.widget = field.hidden_widget()
        if combination == "1:2" or combination == "1:1" or combination == "2:1":
            field = form.fields["trafo_output_bus_3"]
            field.widget = field.hidden_widget()
            field = form.fields["trafo_output_conversionf_3"]
            field.widget = field.hidden_widget()
        if combination == "1:1" or combination == "2:1":
            field = form.fields["trafo_output_bus_2"]
            field.widget = field.hidden_widget()
            field = form.fields["trafo_output_conversionf_2"]
            field.widget = field.hidden_widget()
        if combination == "1:2" or combination == "1:1":
            field = form.fields["trafo_input_bus_2"]
            field.widget = field.hidden_widget()
            field = form.fields["trafo_input_conversionf_2"]
            field.widget = field.hidden_widget()
            field = form.fields["trafo_input_conversionf_1"]
            field.widget = field.hidden_widget()
        if combination == "2:1":
            field = form.fields["trafo_output_conversionf_1"]
            field.widget = field.hidden_widget()


def build_oemof_trafo_expert(list_transformers, k, i, ep_costs):
    if i["trafo_input_output_variation_choice"] == "2:1":
        # print(i)

        input_bus_1 = re.findall(r"\d+", i["trafo_input_bus_1"])
        queryset_input_bus_1 = Bus.objects.filter(id=int(input_bus_1[0]))
        for input_bus_1 in queryset_input_bus_1:
            print(input_bus_1.id, input_bus_1.name, input_bus_1.type)

        input_bus_2 = re.findall(r"\d+", i["trafo_input_bus_2"])
        queryset_input_bus_2 = Bus.objects.filter(id=int(input_bus_2[0]))
        for input_bus_2 in queryset_input_bus_2:
            print(input_bus_2.id, input_bus_2.name, input_bus_2.type)

        output_bus_1 = re.findall(r"\d+", i["trafo_output_bus_1"])
        queryset_output_bus_1 = Bus.objects.filter(id=int(output_bus_1[0]))
        for output_bus_1 in queryset_output_bus_1:
            print(output_bus_1.id, output_bus_1.name, output_bus_1.type)

        print(i["trafo_output_bus_1"] == i["trafo_variableCosts_bus_choice"])

        try:
            list_transformers.append(
                InRetEnsysTransformer(
                    label=i["label"],
                    inputs={
                        input_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_input_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                        input_bus_2.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_input_bus_2"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                    },
                    outputs={
                        output_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        )
                    },
                    conversion_factors={
                        input_bus_1.name: 1 / i["trafo_input_conversionf_1"],
                        input_bus_2.name: 1 / i["trafo_input_conversionf_2"],
                    },
                )
            )

            # print("\nEnergy Conversion (myTransformer): \n")
            # print("{} : {}".format(k, i))

        except Exception as e:
            error_msg = f"Trafo Scenario Serialization ERROR! User: {scenario.project.user.username}. Scenario Id: {scenario.id}. Thrown Exception: {e}."
            logger.error(error_msg)

            raise Exception(error_msg + " - 407")

    elif i["trafo_input_output_variation_choice"] == "2:3":
        # print(i)
        # inputs
        input_bus_1 = re.findall(r"\d+", i["trafo_input_bus_1"])
        queryset_input_bus_1 = Bus.objects.filter(id=int(input_bus_1[0]))
        for input_bus_1 in queryset_input_bus_1:
            print(input_bus_1.id, input_bus_1.name, input_bus_1.type)

        input_bus_2 = re.findall(r"\d+", i["trafo_input_bus_2"])
        queryset_input_bus_2 = Bus.objects.filter(id=int(input_bus_2[0]))
        for input_bus_2 in queryset_input_bus_2:
            print(input_bus_2.id, input_bus_2.name, input_bus_2.type)
        # outputs
        output_bus_1 = re.findall(r"\d+", i["trafo_output_bus_1"])
        queryset_output_bus_1 = Bus.objects.filter(id=int(output_bus_1[0]))
        for output_bus_1 in queryset_output_bus_1:
            print(output_bus_1.id, output_bus_1.name, output_bus_1.type)

        output_bus_2 = re.findall(r"\d+", i["trafo_output_bus_2"])
        queryset_output_bus_2 = Bus.objects.filter(id=int(output_bus_2[0]))
        for output_bus_2 in queryset_output_bus_2:
            print(output_bus_2.id, output_bus_2.name, output_bus_2.type)

        output_bus_3 = re.findall(r"\d+", i["trafo_output_bus_3"])
        queryset_output_bus_3 = Bus.objects.filter(id=int(output_bus_3[0]))
        for output_bus_3 in queryset_output_bus_3:
            print(output_bus_3.id, output_bus_3.name, output_bus_3.type)

        try:
            list_transformers.append(
                InRetEnsysTransformer(
                    label=i["label"],
                    inputs={
                        input_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_input_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                        input_bus_2.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_input_bus_2"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                    },
                    outputs={
                        output_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                        output_bus_2.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 2000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_2"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                        output_bus_3.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 3000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_3"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                    },
                    conversion_factors={
                        input_bus_1.name: i["trafo_input_conversionf_1"],
                        input_bus_2.name: i["trafo_input_conversionf_2"],
                        output_bus_1.name: i["trafo_output_conversionf_1"],
                        output_bus_2.name: i["trafo_output_conversionf_2"],
                        output_bus_3.name: i["trafo_output_conversionf_3"],
                    },
                )
            )

            # print("\nEnergy Conversion (myTransformer): \n")
            # print("{} : {}".format(k, i))

        except Exception as e:
            error_msg = f"Trafo Scenario Serialization ERROR! User: {scenario.project.user.username}. Scenario Id: {scenario.id}. Thrown Exception: {e}."
            logger.error(error_msg)

            raise Exception(error_msg + " - 407")

    elif i["trafo_input_output_variation_choice"] == "3:3":
        # print(i)
        # inputs
        input_bus_1 = re.findall(r"\d+", i["trafo_input_bus_1"])
        queryset_input_bus_1 = Bus.objects.filter(id=int(input_bus_1[0]))
        for input_bus_1 in queryset_input_bus_1:
            print(input_bus_1.id, input_bus_1.name, input_bus_1.type)

        input_bus_2 = re.findall(r"\d+", i["trafo_input_bus_2"])
        queryset_input_bus_2 = Bus.objects.filter(id=int(input_bus_2[0]))
        for input_bus_2 in queryset_input_bus_2:
            print(input_bus_2.id, input_bus_2.name, input_bus_2.type)

        input_bus_3 = re.findall(r"\d+", i["trafo_input_bus_3"])
        queryset_input_bus_3 = Bus.objects.filter(id=int(input_bus_3[0]))
        for input_bus_3 in queryset_input_bus_3:
            print(input_bus_3.id, input_bus_3.name, input_bus_3.type)
        # outputs
        output_bus_1 = re.findall(r"\d+", i["trafo_output_bus_1"])
        queryset_output_bus_1 = Bus.objects.filter(id=int(output_bus_1[0]))
        for output_bus_1 in queryset_output_bus_1:
            print(output_bus_1.id, output_bus_1.name, output_bus_1.type)

        output_bus_2 = re.findall(r"\d+", i["trafo_output_bus_2"])
        queryset_output_bus_2 = Bus.objects.filter(id=int(output_bus_2[0]))
        for output_bus_2 in queryset_output_bus_2:
            print(output_bus_2.id, output_bus_2.name, output_bus_2.type)

        output_bus_3 = re.findall(r"\d+", i["trafo_output_bus_3"])
        queryset_output_bus_3 = Bus.objects.filter(id=int(output_bus_3[0]))
        for output_bus_3 in queryset_output_bus_3:
            print(output_bus_3.id, output_bus_3.name, output_bus_3.type)

        try:
            list_transformers.append(
                InRetEnsysTransformer(
                    label=i["label"],
                    inputs={
                        input_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_input_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                        input_bus_2.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_input_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_input_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_input_bus_2"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                        input_bus_3.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_input_bus_3"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_input_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_input_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_input_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_input_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_input_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_input_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_input_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_input_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_input_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_input_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_input_bus_3"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                    },
                    outputs={
                        output_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                        output_bus_2.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 2000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_2"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                        output_bus_3.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_3"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 3000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_3"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_3"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                    },
                    conversion_factors={
                        input_bus_1.name: i["trafo_input_conversionf_1"],
                        input_bus_2.name: i["trafo_input_conversionf_2"],
                        input_bus_3.name: i["trafo_input_conversionf_3"],
                        output_bus_1.name: i["trafo_output_conversionf_1"],
                        output_bus_2.name: i["trafo_output_conversionf_2"],
                        output_bus_3.name: i["trafo_output_conversionf_3"],
                    },
                )
            )

            # print("\nEnergy Conversion (myTransformer): \n")
            # print("{} : {}".format(k, i))

        except Exception as e:
            error_msg = f"Trafo Scenario Serialization ERROR! User: {scenario.project.user.username}. Scenario Id: {scenario.id}. Thrown Exception: {e}."
            logger.error(error_msg)

            raise Exception(error_msg + " - 407")

    elif i["trafo_input_output_variation_choice"] == "1:1":
        # print(i)

        input_bus_1 = re.findall(r"\d+", i["trafo_input_bus_1"])
        queryset_input_bus_1 = Bus.objects.filter(id=int(input_bus_1[0]))
        for input_bus_1 in queryset_input_bus_1:
            print(input_bus_1.id, input_bus_1.name, input_bus_1.type)

        output_bus_1 = re.findall(r"\d+", i["trafo_output_bus_1"])
        queryset_output_bus_1 = Bus.objects.filter(id=int(output_bus_1[0]))
        for output_bus_1 in queryset_output_bus_1:
            print(output_bus_1.id, output_bus_1.name, output_bus_1.type)

        try:
            list_transformers.append(
                InRetEnsysTransformer(
                    label=i["label"],
                    inputs={
                        input_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_input_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                    },
                    outputs={
                        output_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        )
                    },
                    conversion_factors={
                        output_bus_1.name: i["trafo_output_conversionf_1"]
                    },
                )
            )

            # print("\nEnergy Conversion (myTransformer): \n")
            # print("{} : {}".format(k, i))

        except Exception as e:
            error_msg = f"Trafo Scenario Serialization ERROR! User: {scenario.project.user.username}. Scenario Id: {scenario.id}. Thrown Exception: {e}."
            logger.error(error_msg)

            raise Exception(error_msg + " - 407")

    elif i["trafo_input_output_variation_choice"] == "1:2":
        # print(i)

        input_bus_1 = re.findall(r"\d+", i["trafo_input_bus_1"])
        queryset_input_bus_1 = Bus.objects.filter(id=int(input_bus_1[0]))
        for input_bus_1 in queryset_input_bus_1:
            print(input_bus_1.id, input_bus_1.name, input_bus_1.type)

        output_bus_1 = re.findall(r"\d+", i["trafo_output_bus_1"])
        queryset_output_bus_1 = Bus.objects.filter(id=int(output_bus_1[0]))
        for output_bus_1 in queryset_output_bus_1:
            print(output_bus_1.id, output_bus_1.name, output_bus_1.type)

        output_bus_2 = re.findall(r"\d+", i["trafo_output_bus_2"])
        queryset_output_bus_2 = Bus.objects.filter(id=int(output_bus_2[0]))
        for output_bus_2 in queryset_output_bus_2:
            print(output_bus_2.id, output_bus_2.name, output_bus_2.type)

        try:
            list_transformers.append(
                InRetEnsysTransformer(
                    label=i["label"],
                    inputs={
                        input_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_input_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_input_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_input_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                    },
                    outputs={
                        output_bus_1.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_1"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 1000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_1"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_1"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                        output_bus_2.name: InRetEnsysFlow(
                            # We first assume that it is a base load.
                            # fix=i["input_timeseries"]["value"]
                            # if i["input_timeseries"]["value"]
                            # else None,
                            variable_costs=(
                                i["variable_costs"]["value"]
                                if i["variable_costs"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_variableCosts_bus_choice"]
                                else None
                            ),
                            nominal_value=(
                                i["nominal_value"]["value"]
                                if i["nominal_value"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_max=(
                                i["summed_max"]["value"]
                                if i["summed_max"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            summed_min=(
                                i["summed_min"]["value"]
                                if i["summed_min"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            nonconvex=(
                                InRetEnsysNonConvex()
                                if i["nonconvex"]["value"] == True
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _min=(
                                i["_min"]["value"]
                                if i["_min"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            _max=(
                                i["_max"]["value"]
                                if i["_max"]
                                and i["trafo_output_bus_2"]
                                == i["trafo_technicalp_bus_choice"]
                                else None
                            ),
                            custom_attributes={
                                "renewable_factor": (
                                    i["renewable_factor"]["value"]
                                    if i["renewable_factor"]
                                    else None
                                )
                            },
                            investment=(
                                InRetEnsysInvestment(
                                    ep_costs=ep_costs,
                                    maximum=(
                                        i["maximum"]["value"]
                                        if bool(i["maximum"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 2000000
                                    ),
                                    minimum=(
                                        i["minimum"]["value"]
                                        if bool(i["minimum"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    existing=(
                                        i["existing"]["value"]
                                        if bool(i["existing"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    offset=(
                                        i["offset"]["value"]
                                        if bool(i["offset"])
                                        and i["trafo_output_bus_2"]
                                        == i["trafo_invest_bus_choice"]
                                        else 0
                                    ),
                                    nonconvex=True if bool(i["offset"]) else False,
                                )
                                if bool(ep_costs)
                                and i["trafo_output_bus_2"]
                                == i["trafo_invest_bus_choice"]
                                else None
                            ),
                        ),
                    },
                    conversion_factors={
                        output_bus_1.name: i["trafo_output_conversionf_1"],
                        output_bus_2.name: i["trafo_output_conversionf_2"],
                    },
                )
            )

            # print("\nEnergy Conversion (myTransformer): \n")
            # print("{} : {}".format(k, i))

        except Exception as e:
            error_msg = f"Trafo Scenario Serialization ERROR! User: {scenario.project.user.username}. Scenario Id: {scenario.id}. Thrown Exception: {e}."
            logger.error(error_msg)

            raise Exception(error_msg + " - 407")

    return list_transformers
